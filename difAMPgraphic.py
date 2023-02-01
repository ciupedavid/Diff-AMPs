import time
import zipfile
import json
import shutil
import os
import pywinauto.keyboard
import pywinauto.mouse
import csv
import openpyxl
import pandas as pd
import openpyxl as xl
from openpyxl import load_workbook
from openpyxl.chart import (
    ScatterChart,
    Reference,
    Series
)
from openpyxl.chart.axis import ChartLines

with open(r'Devicee.json') as d:
    paths = json.load(d)['Nimble'][0]

project_location = paths['project_location']
device = paths['device']
gain = paths['gain']

unzip_path_ltspice = project_location + '\\' + device + '\\' + 'LTspice - ' + device + ' G' + gain + '.zip'
print(unzip_path_ltspice)
with zipfile.ZipFile(unzip_path_ltspice) as zip_ref:
    new_path = project_location + '\\' + device
    print(new_path)
    zip_ref.extractall(new_path)

unzip_path_nimble = project_location + '\\' + device + '\\' + 'Nimble - ' + device + ' G' + gain + '.zip'
print(unzip_path_nimble)
with zipfile.ZipFile(unzip_path_nimble) as zip_ref:
    zip_ref.extractall(new_path)

time.sleep(1)

raw_data = project_location + '\\' + device + '\\' + 'Raw Data' + '\\' + 'Individual Stage Data' + '\\' + 'Amplifier' + '\\' + 'Amplifier - Input Referred Noise.csv'
shutil.move(raw_data, project_location + device)

time.sleep(3)

project_loc = paths['project_location']
device = paths['device']
gain = paths['gain']

os.startfile(project_loc + '\\' + device + '\\' + 'AC_Simulation.asc')
time.sleep(1)
pywinauto.keyboard.send_keys("%{S}")
pywinauto.keyboard.send_keys("{R}")
time.sleep(1)
pywinauto.keyboard.send_keys("%{V}")
pywinauto.keyboard.send_keys("{V}")
pywinauto.keyboard.send_keys("{DOWN}")
pywinauto.keyboard.send_keys("{DOWN}")
pywinauto.keyboard.send_keys("{DOWN}")
pywinauto.keyboard.send_keys("{UP}")
pywinauto.keyboard.send_keys("{ENTER}")
time.sleep(1)
pywinauto.keyboard.send_keys("^{TAB}")
pywinauto.keyboard.send_keys("%{F}")
pywinauto.keyboard.send_keys("{E}")
pywinauto.keyboard.send_keys("{ENTER}")
pywinauto.keyboard.send_keys("%{F4}")

with open(r'Devicee.json') as d:
    paths = json.load(d)['Nimble'][0]


input_file = paths['project_location'] + '\\' + paths['device'] + '\\' + 'AC_Simulation.txt'
output_file = paths['project_location'] + '\\' + paths['device'] + '\\' + 'AC_Simulation.xlsx'

wb = openpyxl.Workbook()
ws = wb.worksheets[0]

with open(input_file, 'r') as data:
    reader = csv.reader(data, delimiter='\t')
    for row in reader:
        ws.append(row)

wb.save(output_file)

with open(r'Devicee.json') as d:
    paths = json.load(d)['Nimble'][0]

df = pd.read_excel(paths['project_location'] + '\\' + paths['device'] + '\\' + 'AC_Simulation.xlsx')
df.rename(columns={'V(fb+-1)':'vout'}, inplace=True)
df['vout'] = df['vout'].str.split('(', expand=True)[1]
df['vout'] = df['vout'].str.split('d', expand=True)[0]
df['vout'] = '=VALUE(' + df['vout'] + ')'
df.to_excel(paths['project_location'] + '\\' + paths['device'] + '\\' + 'AC_SimulationEdit.xlsx')

read_file = pd.read_csv(paths['project_location'] + '\\' + paths['device'] + '\\' + 'Amplifier - Input Referred Noise.csv')
read_file.to_excel(paths['project_location'] + '\\' + paths['device'] + '\\' + 'Amplifier - Input Referred Noise.xlsx', index = None, header=True)

with open(r'Devicee.json') as d:
    paths = json.load(d)['Nimble'][0]

path1 = paths['project_location'] + '\\' + paths['device'] + '\\' + 'AC_SimulationEdit.xlsx'
path2 = paths['project_location'] + '\\' + paths['device'] + '\\' + 'Amplifier - Input Referred Noise.xlsx'

wb1 = xl.load_workbook(filename=path1)
ws1 = wb1.worksheets[0]

wb2 = xl.load_workbook(filename=path2)
ws2 = wb2.create_sheet(ws1.title)

for row in ws1:
    for cell in row:
        ws2[cell.coordinate].value = cell.value

wb2.save(path2)

excel_path = paths['project_location'] + '\\' + paths['device'] + '\\' + 'Amplifier - Input Referred Noise.xlsx'

file = openpyxl.load_workbook(excel_path)
sheet_obj = file.active
sheet_obj.delete_cols(2)
sheet_obj.delete_cols(2)
sheet_obj.delete_cols(2)
sheet_obj.delete_cols(2)
sheet_obj.delete_cols(2)
sheet_obj.delete_cols(3)
sheet_obj.delete_cols(3)
sheet_obj.delete_cols(3)

file.save(excel_path)

xl = openpyxl.load_workbook(excel_path)
sheet1 = xl['Sheet11']
sheet2 = xl['Sheet1']

columnA = []

for i in range(1, 1000, 1):
    columnA.append(sheet1.cell(row=i, column=2).value)

for i in range(1, 1000, 1):

    for i in range(1, 1000, 1):
        sheet2.cell(row=i, column=4).value = columnA[i - 1]

columnB = []

for i in range(1, 1000, 1):
    columnB.append(sheet1.cell(row=i, column=3).value)

for i in range(1, 1000, 1):

    for i in range(1, 1000, 1):
        sheet2.cell(row=i, column=5).value = columnB[i - 1]

del xl['Sheet11']

sheet2.cell(row=1, column=1).value = "Frequency (Hz)"
sheet2.cell(row=1, column=2).value = "Total (nV/rt(Hz))'"
sheet2.cell(row=1, column=4).value = "Ltspice Freq"
sheet2.cell(row=1, column=5).value = "Ltspice V(onoise)"

xl.save(excel_path)

with open(r'Devicee.json') as d:
    paths = json.load(d)['Nimble'][0]

workbook = load_workbook(excel_path)
sheet = workbook.active

x_nimble = Reference(sheet, min_col=2, min_row=2, max_row=1000)
y_nimble = Reference(sheet, min_col=1, min_row=2, max_row=1000)
x_ltspice = Reference(sheet, min_col=5, min_row=2, max_row=1000)
y_ltspice = Reference(sheet, min_col=4, min_row=2, max_row=1000)
# x_datasheet = Reference(sheet, min_col=6, min_row=2, max_row=1000)
# y_datasheet = Reference(sheet, min_col=5, min_row=2, max_row=1000)

series_voltage = Series(x_nimble, y_nimble, title_from_data=False, title="Nimble")
series_freq = Series(x_ltspice, y_ltspice, title_from_data=False, title="LTspice")

# Chart type
chart = ScatterChart()
chart.series.append(series_voltage)
chart.series.append(series_freq)
# chart.series.append(series_mag)
# chart.x_axis.scaling.logBase = 10
# chart.y_axis.scaling.logBase = 10
# chart.x_axis.number_format = '0.00E+00'
# chart.x_axis.minorGridlines = ChartLines()
# chart.y_axis.minorGridlines = ChartLines()

chart.x_axis.scaling.min = paths['x_axis_min']
chart.y_axis.scaling.min = paths['y_axis_min']
chart.x_axis.scaling.max = paths['x_axis_max']
chart.y_axis.scaling.max = paths['y_axis_max']
chart.x_axis.tickLblPos = "low"
# chart.x_axis.tickLblSkip = 3

chart.title = None
chart.x_axis.title = 'Frequency (Hz)'
chart.y_axis.title = 'Voltage Noise Density (nV√Hz)'
chart.legend.position = 'r'

sheet.add_chart(chart, 'K3')
workbook.save(excel_path)

extra_files_remove = paths['project_location'] + paths['device']
zip_remove_nimble = extra_files_remove + '\\' + 'Nimble - ' + paths['device'] + ' G' + paths['gain'] + '.zip'
zip_remove_ltspice = extra_files_remove + '\\' + 'LTspice - ' + paths['device'] + ' G' + paths['gain'] + '.zip'
os.remove(extra_files_remove + '\\' + 'AC_Simulation.asc')
os.remove(extra_files_remove + '\\' + 'Noise_Simulation.asc')
os.remove(extra_files_remove + '\\' + 'Transient_Simulation.asc')
os.remove(extra_files_remove + '\\' + 'AC_Simulation.log')
os.remove(extra_files_remove + '\\' + 'AC_Simulation.raw')
os.remove(extra_files_remove + '\\' + 'AC_Simulation.txt')
os.remove(extra_files_remove + '\\' + 'AC_Simulation.xlsx')
os.remove(extra_files_remove + '\\' + 'AC_SimulationEdit.xlsx')
os.remove(extra_files_remove + '\\' + 'AC_Simulation.op.raw')
os.remove(extra_files_remove + '\\' + 'Amplifier - Input Referred Noise.csv')
os.remove(zip_remove_nimble)
os.remove(zip_remove_ltspice)
shutil.rmtree(extra_files_remove + '\\' + 'Raw Data')